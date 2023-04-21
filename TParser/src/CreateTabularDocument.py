import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# Create a new workbook
# Global variables that we are going to include 

_border_style_bottom = Border(bottom=Side(style='thin'))
_border_style_left = Border(left=Side(style='thin'))
_border_style_right = Border(right=Side(style='thin'))

def createDocument(fileNameInput, studentInformation, upperDivisionCourses, lowerDivisionCourses, technicalCSECourses, mathRequiredCourses, scienceCourses, sbcCourses, classesPerSemester):

    # We might not do threads, so we have to see 
    workbook = openpyxl.Workbook()

    sheet = workbook.create_sheet()
    ws = workbook['Sheet']
    workbook.remove(ws)

    starterColumnForPlanningGrid = createStudentTableInformation(sheet, studentInformation)
    starterColumnForMajorReqs = createPlanningGrid(sheet, starterColumnForPlanningGrid, classesPerSemester)
    createCSEMajorRequirements(sheet, starterColumnForMajorReqs, studentInformation, upperDivisionCourses, lowerDivisionCourses, technicalCSECourses, mathRequiredCourses, scienceCourses, sbcCourses)

    fileName = fileNameInput + '.xlsx'
    workbook.save(fileName)
    return fileName

@staticmethod
def createStudentTableInformation(sheet, studentInformation):

    row_value_start = 1 # Value subject to change, depending on user 
    start_col = 1 # Subject to change
    end_col = 2 # Subject to change, depends on the user 

    column_counter = start_col # Subject to change 

    sheet.append(['Student Information'])
    sheet.merge_cells(start_row=row_value_start, start_column=start_col, end_row=row_value_start, end_column=end_col)
    while column_counter != end_col + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=start_col).font = Font(size=14, bold=True, name='Calibri')

    # Add data to the Student Information table
    # This needs to be manually changed if it is required to add more fields
    data = [
        ['Last Updated'],
        ['Name'],
        ['ID Number'],
        ['Requirement Term'],
        ['CSHP'],
        ['Cumulative GPA'],
        ['Cumulative Credits'],
        ['Upper Division Credits']
    ]

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20

    # Add data to the table

    rowTrackerForData = 2
    for row in data:
        sheet.append(row)
        getKey = row.pop(0)
        getStudentInfo = studentInformation[getKey]
        try:
            sheet.cell(row=rowTrackerForData, column=end_col).value = float(getStudentInfo)
            sheet.cell(row=rowTrackerForData, column=end_col).alignment = Alignment(horizontal='left')
        except:
            sheet.cell(row=rowTrackerForData, column=end_col).value = str(getStudentInfo)
            sheet.cell(row=rowTrackerForData, column=end_col).alignment = Alignment(horizontal='left')
        rowTrackerForData += 1

    # Apply formatting to the Student Information table
    end_row_value = 0
    for i in range(1, 11):
        end_row_value += 1
        sheet.cell(row=i, column=start_col).font = Font(bold=True)
        sheet.cell(row=i, column=start_col).alignment = Alignment(horizontal='left')
    sheet.cell(row=row_value_start, column=start_col).font = Font(size=14, bold=True, name='Calibri')
    end_row_value -= 1 # Decrease by 1 since loop will run an extra time

    # Add Border To The Table 

    bottom_and_left_border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

    sheet.cell(row=row_value_start, column=start_col).border = _border_style_right
    sheet.cell(row=end_row_value, column=start_col).border = _border_style_bottom
    sheet.cell(row=end_row_value, column=end_col).border = bottom_and_left_border

    counter = 2
    for i in range(counter, 9):
        sheet.cell(row=i, column=end_col).border = _border_style_right
    return end_col+2 # We will start at a second col
        

# A dynamic approach to this such that we add to the planning grid based on the number the information we get from the dictionary
@staticmethod
def createPlanningGrid(sheet, starter_column, classesPerSemester):

    row_value_start = 1 # Subject to change 
    col_value_start = starter_column # Subject to change
    col_value_end = 11 # Subject to change

    # For the planning grid, we need this to be a length of 8 no matter what


    # Get the last index of the list
    last_key = list(classesPerSemester.keys())[-1]

    acceptable_grades = ('A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'XFER')

    # Used for purposefully dynamically adjusting the table, does not affect original information 
    if len(classesPerSemester) % 4 != 0: # Create filler dates and aff them to schedule
        fakeClassesAddedCounter = 0

        priorityTerm = {'Fall': 'Winter', 'Winter': 'Spring', 'Spring': 'Summer', 'Summer': 'Fall'}

        while len(classesPerSemester) % 4 != 0:
            keyAtLastIndex = ""
            keyCounter = 0
            for keys in classesPerSemester.keys():
                if keyCounter == len(classesPerSemester) - 1:
                    keyAtLastIndex = keys
                keyCounter += 1
            # Once we get the key at the last index, we then want to parse this information 
            whiteSpaceIndex = keyAtLastIndex.index(' ')
            semesterTerm = keyAtLastIndex[0:whiteSpaceIndex]
            semesterYear = keyAtLastIndex[whiteSpaceIndex+1:]

            getNewTerm = priorityTerm[semesterTerm]
            if getNewTerm == 'Winter': semesterYear = float(semesterYear) + 1 # Add to the semester Year
            listOfClassesInFakeSemester = []
            classesPerSemester[getNewTerm + " " + str(semesterYear)] = listOfClassesInFakeSemester
            fakeClassesAddedCounter += 1
        
    column_counter = col_value_start
    sheet.cell(row=row_value_start, column=col_value_start).value = "Courses Taken & Credits Earned"
    sheet.merge_cells(start_row=row_value_start, start_column = col_value_start, end_row=row_value_start, end_column=col_value_end)
    while column_counter != col_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=col_value_start).font = Font(size=14, bold=True, name='Calibri')


    highestListValue = 0
    for keys in classesPerSemester:
        getList = classesPerSemester[keys]
        lengthOfList = len(getList)
        if lengthOfList > highestListValue:
            highestListValue = lengthOfList
    
    highestListValue = highestListValue + 1 

    starting_row = row_value_start + 1 # This only gets incremented when we hit starting_column value of 11
    starting_column = col_value_start # We increment this until we hit the value of K (which is the value of 11 )

    keys_list = list(classesPerSemester.keys())
    lengthOfListKeys = len(keys_list)
    index = 0

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    recentClass = False
    while lengthOfListKeys != 0:

        getFirstIndex = keys_list[index]
        if getFirstIndex == last_key: recentClass = True
        index += 1
        indexOfWhiteSpace = getFirstIndex.index(' ')
        seasonInformation = getFirstIndex[0:indexOfWhiteSpace]
        seasonYear = getFirstIndex[indexOfWhiteSpace+1:]

        if starting_column < col_value_end: 
            sheet.cell(row=starting_row, column=starting_column).value = seasonInformation
            sheet.cell(row=starting_row, column=starting_column).border = all_sides_border
            try: sheet.cell(row=starting_row, column=starting_column+1).value = float(seasonYear) 
            except: sheet.cell(row=starting_row, column=starting_column+1).value = str(seasonYear) 
            sheet.cell(row=starting_row, column=starting_column+1).border = all_sides_border

            # After we added the information, we now add the classes at the bottom

            getList = classesPerSemester[seasonInformation + " " + seasonYear]
            tempStartingRow = starting_row
            tempStartingCol = starting_column
            classCounter = 0
            creditsCounter = 0
            while classCounter != highestListValue:
                try:
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).value = getList[0].courseName if (recentClass != True and getList[0].grade != 'I') else getList[0].courseName + "*"
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).value =  float(getList[0].credits) if (getList[0].grade in acceptable_grades) else (float(0) if str(getList[0].grade) == "" else str(getList[0].grade))
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).alignment = Alignment(horizontal='right')
                    creditsCounter += float(getList[0].credits)
                    getList.pop(0)
                    tempStartingRow += 1
                    classCounter += 1
                except:
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    tempStartingRow += 1
                    classCounter += 1
            # Reached here when we are done filling out the classes, so now we need to put the total credits

            # We use the highestListValue for this 

            sheet.cell(row=starting_row+highestListValue, column=starting_column).value = "Total: "
            sheet.cell(row=starting_row+highestListValue, column=starting_column).border = all_sides_border
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).value = creditsCounter
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).border = all_sides_border

            #Make bold this part
            sheet.cell(row=starting_row+highestListValue, column=starting_column).font = Font(bold=True)
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).font = Font(bold=True)
            
            starting_column += 2
            lengthOfListKeys -= 1

        else: # This is the case when the starting column is greater than 11, which means we need to change our starting row value

            starting_row = starting_row + highestListValue + 1 # The next available space     
            starting_column = 4 # Put the starting column back to 4 

            sheet.cell(row=starting_row, column=starting_column).value = seasonInformation
            sheet.cell(row=starting_row, column=starting_column).border = all_sides_border
            sheet.cell(row=starting_row, column=starting_column+1).value = float(seasonYear)
            sheet.cell(row=starting_row, column=starting_column+1).border = all_sides_border

            getList = classesPerSemester[seasonInformation + " " + seasonYear]
            tempStartingRow = starting_row
            tempStartingCol = starting_column
            classCounter = 0
            creditsCounter = 0
            while classCounter != highestListValue:
                try:
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).value = getList[0].courseName if (recentClass != True and getList[0].grade != 'I') else getList[0].courseName + "*"
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).value = float(getList[0].credits) if (getList[0].grade in acceptable_grades) else (float(0) if str(getList[0].grade) == "" else str(getList[0].grade))
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).alignment = Alignment(horizontal='right')
                    creditsCounter += float(getList[0].credits)
                    getList.pop(0)
                    tempStartingRow += 1
                    classCounter += 1
                except:
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol).border = _border_style_left
                    sheet.cell(row=tempStartingRow+1, column=tempStartingCol+1).border = _border_style_right
                    tempStartingRow += 1
                    classCounter += 1
            # Reached here when we are done filling out the classes, so now we need to put the total credits

            # We use the highestListValue for this 

            sheet.cell(row=starting_row+highestListValue, column=starting_column).value = "Total: "
            sheet.cell(row=starting_row+highestListValue, column=starting_column).border = all_sides_border
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).value = creditsCounter
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).border = all_sides_border

            sheet.cell(row=starting_row+highestListValue, column=starting_column).font = Font(bold=True)
            sheet.cell(row=starting_row+highestListValue, column=starting_column+1).font = Font(bold=True)

            starting_column += 2
            lengthOfListKeys -= 1
    
    sheet.merge_cells(start_row=starting_row+highestListValue+1, start_column = col_value_start, end_row=starting_row+highestListValue+1, end_column=col_value_end)
    sheet.cell(row=starting_row+highestListValue+1, column=col_value_start).value = "* Means Course Is In-Progress (includes Incomplete courses)"
    sheet.cell(row=starting_row+highestListValue+1, column=col_value_start).font = Font(bold=True)
    return col_value_end+2 # 2 columns apart space 

# Same process: we want to dynamically modify this table to reflect certain conditions 
@staticmethod
def createCSEMajorRequirements(sheet, starter_column, studentInformation, upperDivisionCourses, lowerDivisionCourses, technicalCSECourses, mathRequiredCourses, scienceCourses, sbcCourses):

    row_value_start = 1 # Subject to change
    column_value_start = starter_column # Subject to change
    column_value_end = 21 # Subject to change 

    if (column_value_end - column_value_start + 1) % 3 != 0:
        column_value_start += 2 # Shift by 2 
        column_value_end += 3


    column_counter = column_value_start

    sheet.cell(row=row_value_start, column=column_value_start).value = "CSE BS Degree Requirements"
    sheet.merge_cells(start_row=row_value_start, start_column = column_value_start, end_row=row_value_start, end_column=column_value_end)
    while column_counter != column_value_end + 1:
        sheet.cell(row=row_value_start, column=column_counter).font = Font(bold=True)
        sheet.cell(row=row_value_start, column=column_counter).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_value_start, column=column_counter).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        column_counter += 1
    sheet.cell(row=row_value_start, column=column_value_start).font = Font(size=14, bold=True, name='Calibri')

    # NO MULTI THREADING WILL BE USED FOR SPECIFIC REASONS 

    upperDivisionCSERowStart = handleLowerDivisonCourses(sheet, row_value_start, column_value_start, column_value_end, lowerDivisionCourses) # Method Complete!
    technicalCSERowStart = handleUpperDivisionCourses(sheet, upperDivisionCSERowStart, column_value_start, column_value_end, studentInformation, upperDivisionCourses) # Method Complete!
    mathCoursesRowStart = handleTechnicalCourses(sheet, technicalCSERowStart, column_value_start, column_value_end, technicalCSECourses) # Method Works!
    scienceCoursesRowStart = handleMathRequiredCourses(sheet, mathCoursesRowStart, column_value_start, column_value_end, mathRequiredCourses) # Method Works!
    sbcCoursesRowStart = handleScienceRequiredCourses(sheet, scienceCoursesRowStart, column_value_start, column_value_end, scienceCourses) # Method Works!
    handleSBCsCourses(sheet, sbcCoursesRowStart, column_value_start, column_value_end, sbcCourses)


@staticmethod
def handleLowerDivisonCourses(sheet, row_value_start, column_value_start, column_value_end, lowerDivisionCourses):

    valuesToPopAndInsert = {'Lower Division Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}

    row_start = row_value_start + 1
    row_limit = 5

    lower_division_data = ['CSE 114 (TECH)', 'CSE 214', 'CSE 215', 'CSE 216', 'CSE 220']

    returnNextRowStarter = excelFileInputterAlgorithm(sheet, row_start, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, lower_division_data, lowerDivisionCourses)
    return returnNextRowStarter


@staticmethod
def handleUpperDivisionCourses(sheet, upper_start, column_value_start, column_value_end, studentInformation, upperDivisionCourses):

    valuesToPopAndInsert = {'Upper Division Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}
    
    row_limit = 0
    getCSHPValue = studentInformation['CSHP']
    if getCSHPValue == "NO": row_limit = 8
    else: row_limit = 10
    
    upper_division_data = ['CSE 300 (SPK/WRTD)', 'CSE 303', 'CSE 310', 'CSE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE 373', 'CSE 416 (ESI/EXP+/SBS+/STEM+)'] if getCSHPValue == "NO" else ['CSE 300 (SPK/WRTD)', 'CSE 303', 'CSE 310', 'CSE 312 (CER/ESI/STAS)', 'CSE 316 (ESI/EXP+/SBS+/STEM+)', 'CSE 320', 'CSE 373', 'CSE 416 (ESI/EXP+/SBS+/STEM+)', 'CSE 495', 'CSE 496']

    returnNextSectionRowStarter = excelFileInputterAlgorithm(sheet, upper_start, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, upper_division_data, upperDivisionCourses)
    return returnNextSectionRowStarter

  
@staticmethod
def handleTechnicalCourses(sheet, technical_row_starter, column_value_start, column_value_end, technicalCSECourses):

    valuesToPopAndInsert = {'Technical Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}
    
    row_limit = 0
    if len(technicalCSECourses) <= 4:
        row_limit = 4
    elif len(technicalCSECourses) > 4:
        row_limit = len(technicalCSECourses)

    nextSectionRowStarter = excelFileInputterAlgorithm(sheet, technical_row_starter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, [], technicalCSECourses)
    return nextSectionRowStarter

@staticmethod
def handleMathRequiredCourses(sheet, math_start, column_value_start, column_value_end, mathRequiredCourses):

    requiredCourses = ('MAT 125', 'MAT 126', 'MAT 127', 'MAT 131', 'MAT 132', 'MAT 151', 'AMS 161', 'AMS 171', 'MAT 141', 'MAT 142', 'AMS 210', 'MAT 211', 'AMS 301', 'AMS 310', 'AMS 311')

    math_copy = mathRequiredCourses.copy()

    for keys in math_copy: # Cleanse the dictionary 
        if keys not in requiredCourses:
            del mathRequiredCourses[keys] 
    
    valuesToPopAndInsert = {'Math Required Courses': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}

    row_limit = 0
    if '125' in mathRequiredCourses: # This implies that the student may also take '126' and '127' combination
        row_limit = 7
    elif '171' in mathRequiredCourses:
        row_limit = 4
    else:
        row_limit = 5

    nextSectionRowStarter = excelFileInputterAlgorithm(sheet, math_start, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, [], mathRequiredCourses)
    return nextSectionRowStarter

@staticmethod
def handleScienceRequiredCourses(sheet, row_starter, column_value_start, column_value_end, scienceCourses):

    valuesToPopAndInsert = {'Science Courses (9 Credits, 1 Lab)': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}

    labCombos = {'BIO 201': 'BIO 204', 'BIO 202': 'BIO 204', 'BIO 203': 'BIO 204', 'CHE 131': 'CHE 133', 'CHE 152': 'CHE 154', 'PHY 126': 'PHY 133', 'PHY 131': 'PHY 133', 'PHY 141': 'PHY 133'}
    independentCourses = ('AST 203', 'AST 205', 'CHE 132', 'CHE 321', 'CHE 322', 'CHE 331', 'CHE 332', 'GEO 102', 'GEO 103', 'GEO 112', 'GEO 113', 'GEO 122', 'PHY 125', 'PHY 127', 'PHY 132', 'PHY 134', 'PHY 142', 'PHY 251', 'PHY 252')

    science_copy = scienceCourses.copy()
    labClassFound = []
    laboratoryComboFound = False 
    for keys in science_copy:
        if keys in labCombos:
            labClassFound.append(keys)
            labDescription = labCombos[keys] # The class that we need
            list_of_keys = list(scienceCourses.keys())
            if laboratoryComboFound == False:
                for originalKeys in list_of_keys:
                    if labDescription in originalKeys:
                        laboratoryComboFound = True 
                        break
            if laboratoryComboFound == True:
                break # We found a class/lab combo, so we break out of the loop and do nothing 
    
    if laboratoryComboFound == False: # We found the class, but couldn't find the associated lab combo
        for keys in scienceCourses:
            try: 
                retrieveClass = labClassFound[0] 
                labClass = labCombos[retrieveClass]
                if retrieveClass in keys:
                    scienceCourses[keys].courseName = f'{retrieveClass} ({labClass} REQUIRED.)'
                labClassFound.pop(0)
            except: pass

        
    # Clear out any extraneous classes we may have 

    for keys in science_copy:
        if keys not in independentCourses:
            if (keys in labCombos) or (keys in labCombos.values()): pass # We then check the other dictionary pass
            else: del scienceCourses[keys]
    
    # Now we have three cases for the row limit:
    # 1) After iterating through the number of credits, if the sum isn't 9, the default row_limit is 4
    # 2) If the number of credits exceeds or is greater than 9, the default row_limit is how many times the iteration went
    # 3) If the number of credits exceeds or is greater than 9, but no lab combo, then the row limit is going to be number of iterations + 1

    # Go through credits
    creditsCounter = 0
    numberOfIterations = 0
    for keys in scienceCourses:
        creditsCounter += float(scienceCourses[keys].credits)
        numberOfIterations += 1

    row_limit = 0
    if laboratoryComboFound == False and creditsCounter >= 9: row_limit = numberOfIterations + 1
    elif laboratoryComboFound == True and creditsCounter >= 9: row_limit = numberOfIterations
    elif creditsCounter < 9: row_limit = 4 # Even if the lab is found or not found, if the credits is less than 4, then the row_limit is set to 4

    nextSectionRowStarter = excelFileInputterAlgorithm(sheet, row_starter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, [], scienceCourses)
    return nextSectionRowStarter

@staticmethod
def handleSBCsCourses(sheet, row_starter, column_value_start, column_value_end, sbcCourses):

    valuesToPopAndInsert = {'SBC Objectives': 3, 'Grade': 1, 'Credits': 1, 'Term': 1, 'Year': 1, 'Comments': 2}
    sbc_data = ['ARTS', 'GLO', 'HUM', 'SBS', 'USA', 'WRT', 'DIV']
    row_limit = 7 # Only 7 SBCs to add 
    sbcInputterAlgorithm(sheet, row_starter, column_value_start, column_value_end, row_limit, valuesToPopAndInsert, sbc_data, sbcCourses)

        
@staticmethod
def excelFileInputterAlgorithm(sheet, row_value, column_value_start, column_value_end, row_limit, insert_dictionary, class_data, category_dictionary):

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_start+2) # For Lower Division Courses
    sheet.merge_cells(start_row=row_value, start_column=column_value_end-1, end_row=row_value, end_column=column_value_end) # For Comments

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    initial_column = column_value_start 
    keys_in_dict = list(insert_dictionary.keys())
    while len(insert_dictionary) != 0:
        sheet.cell(row=row_value, column=initial_column).value = keys_in_dict[0]
        sheet.cell(row=row_value, column=initial_column).font = Font(size=10, name='Calibri', bold=True)
        valueOffset = insert_dictionary[keys_in_dict[0]]
        del insert_dictionary[keys_in_dict[0]]
        keys_in_dict.pop(0)
        initial_column += valueOffset
    
    initial_column = column_value_start
    while initial_column != column_value_end + 1:
        sheet.cell(row=row_value, column=initial_column).border = all_sides_border
        initial_column += 1
    
    row_lower = row_value + 1
    row_max = row_lower + row_limit

    keys_of_courses = []
    if len(class_data) == 0: # Means no preassumed data
        keys_of_courses = list(category_dictionary.keys())
    else: # There is preassumed data 
        keys_of_courses = list(class_data)

    while row_lower != row_max:
        sheet.merge_cells(start_row=row_lower, start_column=column_value_start, end_row=row_lower, end_column=column_value_start+2)
        sheet.merge_cells(start_row=row_lower, start_column=column_value_end-1, end_row=row_lower, end_column=column_value_end)
        try:
            keyFirst = keys_of_courses[0]
            courseName = ""
            if len(class_data) == 0:
                courseName = category_dictionary[keyFirst].courseName # Injection
            else: # Means there is some pre assumed data
                courseName = keyFirst # Assumption 
            sheet.cell(row=row_lower, column=column_value_start).value = courseName
        except: pass
        sheet.cell(row=row_lower, column=column_value_start).font = Font(size=10)
        sheet.cell(row=row_lower, column=column_value_start).border = _border_style_left # Border to Lower Division
        sheet.cell(row=row_lower, column=column_value_start+3).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+4).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+5).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+6).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end-1).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end).border = _border_style_right
        try:
            keys_of_courses.pop(0)
        except:
            pass
        row_lower += 1
    
    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_lower, column=start_col_merge).border = all_sides_border
        start_col_merge += 1

    # Begin inserting data 

    getListOfSortedKeys = list(sorted(category_dictionary.keys()))
    lengthOfList = len(getListOfSortedKeys)
    creditsCounter = 0
    creditsValue = 0

    searching_start_row = row_value + 1
    row_offset = 1

    while lengthOfList != 0:
        getFirstIndex = getListOfSortedKeys[0]
        if searching_start_row >= row_max: # Not found, so we pop this extraneous key from the list and continue 
            searching_start_row = row_value + 1
            row_offset = 1
            getListOfSortedKeys.pop(0)
            lengthOfList -= 1
        cellValue = sheet.cell(row=searching_start_row, column=column_value_start).value
        if getFirstIndex in cellValue: # Means the courses dictionary has the class we need to parse

            insertion_start = column_value_start + 3

            while insertion_start != column_value_end:
                getColumnDescription = str(sheet.cell(row=searching_start_row-row_offset, column=insertion_start).value)
                getColumnDescription = getColumnDescription.lower()
                classObject = category_dictionary[getFirstIndex]
                field_value = getattr(classObject, getColumnDescription)
                try: 
                    sheet.cell(row=searching_start_row, column=insertion_start).value = float(field_value) # Put the grade of the object
                    if getColumnDescription == "credits":
                        creditsCounter += float(field_value)
                        # We are also gonna wanna get the grade at the previous cell
                        getGradeLetter = str(sheet.cell(row=searching_start_row, column=insertion_start-1).value)
                        if getGradeLetter in acceptable_grades: getPoints = acceptable_grades[getGradeLetter]
                        else: getPoints = 0.00
                        letterAndCreditProduct = float(getPoints) * float(field_value)
                        creditsValue += letterAndCreditProduct
                except: 
                        sheet.cell(row=searching_start_row, column=insertion_start).value = str(field_value) # Put the grade of the object 
                sheet.cell(row=searching_start_row, column=insertion_start).alignment = Alignment(horizontal='left')
                insertion_start += 1

            searching_start_row = row_value + 1 # Reset back to original spot 
            row_offset = 1
            getListOfSortedKeys.pop(0)
            lengthOfList -= 1

        else:
            searching_start_row += 1 # Increment by 1 
            row_offset += 1

    if creditsCounter != 0: totalGPA = round(float(creditsValue/creditsCounter), 3)
    else: totalGPA = round(float(0), 3)

    sheet.merge_cells(start_row=row_max, start_column=column_value_start, end_row=row_max, end_column=column_value_end)
    sheet.cell(row=row_max, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_max, column=column_value_start).value = f'Category GPA: {totalGPA}'
    return row_max+1 # Cause we need this information to be passed down and is also the next row that will hold the section information

@staticmethod
def sbcInputterAlgorithm(sheet, row_value, column_value_start, column_value_end, row_limit, insert_dictionary, class_data, category_dictionary):

    acceptable_grades = {'A': 4.00, 'A-': 3.67, 'B+': 3.33, 'B': 3.00, 'B-': 2.67, 'C+': 2.33, 'C': 2.00, 'C-': 1.67, 'D+': 1.33, 'D': 1.00, '': 0.00}

    sheet.merge_cells(start_row=row_value, start_column=column_value_start, end_row=row_value, end_column=column_value_start+2) # For Lower Division Courses
    sheet.merge_cells(start_row=row_value, start_column=column_value_end-1, end_row=row_value, end_column=column_value_end) # For Comments

    all_sides_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

    initial_column = column_value_start 
    keys_in_dict = list(insert_dictionary.keys())
    while len(insert_dictionary) != 0:
        sheet.cell(row=row_value, column=initial_column).value = keys_in_dict[0]
        sheet.cell(row=row_value, column=initial_column).font = Font(size=10, name='Calibri', bold=True)
        valueOffset = insert_dictionary[keys_in_dict[0]]
        del insert_dictionary[keys_in_dict[0]]
        keys_in_dict.pop(0)
        initial_column += valueOffset
    
    initial_column = column_value_start
    while initial_column != column_value_end + 1:
        sheet.cell(row=row_value, column=initial_column).border = all_sides_border
        initial_column += 1
    
    row_lower = row_value + 1
    row_max = row_lower + row_limit

    keys_of_courses = []
    if len(class_data) == 0: # Means no preassumed data
        keys_of_courses = list(category_dictionary.keys())
    else: # There is preassumed data 
        keys_of_courses = list(class_data)

    while row_lower != row_max:
        sheet.merge_cells(start_row=row_lower, start_column=column_value_start, end_row=row_lower, end_column=column_value_start+2)
        sheet.merge_cells(start_row=row_lower, start_column=column_value_end-1, end_row=row_lower, end_column=column_value_end)
        try:
            keyFirst = keys_of_courses[0]
            courseName = ""
            if len(class_data) == 0:
                courseName = category_dictionary[keyFirst].courseName # Injection
            else: # Means there is some pre assumed data
                courseName = keyFirst # Assumption 
            sheet.cell(row=row_lower, column=column_value_start).value = courseName
        except: pass
        sheet.cell(row=row_lower, column=column_value_start).font = Font(size=10)
        sheet.cell(row=row_lower, column=column_value_start).border = _border_style_left # Border to Lower Division
        sheet.cell(row=row_lower, column=column_value_start+3).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+4).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+5).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_start+6).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end-1).border = _border_style_left
        sheet.cell(row=row_lower, column=column_value_end).border = _border_style_right
        try:
            keys_of_courses.pop(0)
        except:
            pass
        row_lower += 1
    
    start_col_merge = column_value_start
    while start_col_merge != column_value_end + 1:
        sheet.cell(row=row_lower, column=start_col_merge).border = all_sides_border
        start_col_merge += 1


    getListOfSortedKeys = list(category_dictionary.keys())
    lengthOfList = len(getListOfSortedKeys)
    creditsCounter = 0
    creditsValue = 0

    searching_start_row = row_value + 1
    row_offset = 1

    # This algorithm varies in a sense that the object contains a list of SBCs, so we need to insert them based on highest value 

    while lengthOfList != 0:    

        getFirstIndex = getListOfSortedKeys[0]
        getSBCClassList = category_dictionary[getFirstIndex].sbc # Will give us the sbcList

        if searching_start_row >= row_max: # Not found, so we pop this extraneous key from the list and continue 
            getListOfSortedKeys.pop(0) # Will pop the object even if there is an sbc that exists but is not in the required ones 
            searching_start_row = row_value + 1
            row_offset = 1
            lengthOfList -= 1

        cellValue = sheet.cell(row=searching_start_row, column=column_value_start).value

        if cellValue in getSBCClassList: # We found a course that has an sbc

            indexOfCellValue = category_dictionary[getFirstIndex].sbc.index(cellValue)
            sheet.cell(row=searching_start_row, column=column_value_start).value = f'{sheet.cell(row=searching_start_row, column=column_value_start).value} : {category_dictionary[getFirstIndex].courseName}'
            category_dictionary[getFirstIndex].sbc.pop(indexOfCellValue) 

            insertion_start = column_value_start + 3

            while insertion_start != column_value_end:
                getColumnDescription = str(sheet.cell(row=searching_start_row-row_offset, column=insertion_start).value)
                getColumnDescription = getColumnDescription.lower()
                classObject = category_dictionary[getFirstIndex]
                field_value = getattr(classObject, getColumnDescription)
                try: 
                    sheet.cell(row=searching_start_row, column=insertion_start).value = float(field_value) # Put the grade of the object
                    if getColumnDescription == "credits":
                        creditsCounter += float(field_value)
                        # We are also gonna wanna get the grade at the previous cell
                        getGradeLetter = str(sheet.cell(row=searching_start_row, column=insertion_start-1).value)
                        if getGradeLetter in acceptable_grades: getPoints = acceptable_grades[getGradeLetter]
                        else: getPoints = 0.00
                        letterAndCreditProduct = float(getPoints) * float(field_value)
                        creditsValue += letterAndCreditProduct
                except: 
                        sheet.cell(row=searching_start_row, column=insertion_start).value = str(field_value) # Put the grade of the object 
                sheet.cell(row=searching_start_row, column=insertion_start).alignment = Alignment(horizontal='left')
                insertion_start += 1

            if len(category_dictionary[getFirstIndex].sbc) == 0:
                getListOfSortedKeys.pop(0)
                lengthOfList -= 1
        else:
            searching_start_row += 1 # Increment by 1 
            row_offset += 1

    if creditsCounter != 0: totalGPA = round(float(creditsValue/creditsCounter), 3)
    else: totalGPA = round(float(0), 3)

    sheet.merge_cells(start_row=row_max, start_column=column_value_start, end_row=row_max, end_column=column_value_end)
    sheet.cell(row=row_max, column=column_value_start).font = Font(size=10, name='Calibri', bold=True)
    sheet.cell(row=row_max, column=column_value_start).value = f'Category GPA: {totalGPA}'